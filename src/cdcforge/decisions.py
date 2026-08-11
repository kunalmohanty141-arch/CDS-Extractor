"""The decision sheet — a batch of choices a human makes, in Excel.

The tool can rank views, measure coverage and say what it would do. It cannot
know that VBAP is only wanted for one document category, that EKPO is already
covered by a feed someone built last year, or that this quarter only the
finance tables matter. Those are the customer's facts, and they arrive in a
spreadsheet because that is where this industry keeps its lists.

So the batch flow is a round trip. ``plan`` assesses a list of tables and
writes a sheet with a suggestion in every row; the user edits the ``Action``
and ``Base`` columns — in Excel, offline, in a meeting, with colleagues — and
``apply`` reads it back and does what it says.

Two things this module is careful about, both learned from what goes wrong with
spreadsheets rather than with code:

**The suggestion and the decision are different columns.** ``Suggested action``
is read-only provenance; ``Action`` is what will happen. Merging them into one
editable cell loses the record of what the tool thought, and the first question
anyone asks about a batch six weeks later is "why did we do that one".

**A sheet that cannot be executed is rejected whole, not halfway.** Every row is
validated before any of them runs, because a batch that creates eleven objects
and then stops on a typo in row twelve leaves the user to work out which eleven.
"""

from __future__ import annotations

import re
from dataclasses import dataclass, field
from pathlib import Path

#: What a row can ask for.
USE = "USE"
WRAP = "WRAP"
BUILD = "BUILD"
SKIP = "SKIP"

ACTIONS = (USE, WRAP, BUILD, SKIP)

ACTION_HELP = {
    USE: "Replicate this view as it is — it already carries extraction and CDC. "
         "Nothing is generated.",
    WRAP: "Generate a Z wrapper over the view in Base, adding the extraction "
          "and CDC annotations it lacks.",
    BUILD: "Generate a Z view directly over the table. Use when no existing "
           "view is suitable, or when you need every column and every row.",
    SKIP: "Do nothing with this object.",
}

_CUSTOMER_NAME = re.compile(r"^[YZ][A-Z0-9_]{0,29}$")

#: The editable columns, in sheet order. Everything else is provenance.
_EDITABLE = ("Action", "Base", "Target name", "Note")

_COLUMNS = (
    "Object",
    "Kind",
    "Action",
    "Base",
    "Target name",
    "Note",
    "Existing",
    "Suggested action",
    "Suggested base",
    "Coverage",
    "Why",
    "Candidates",
)


class PlanError(Exception):
    """The sheet cannot be executed as written."""


@dataclass
class Decision:
    """One row: an object, and what to do about it."""

    object_name: str
    kind: str = "TABLE"
    action: str = SKIP
    base: str = ""
    target: str = ""
    note: str = ""

    # -- provenance, written by `plan` and never read back as instruction ---
    suggested_action: str = ""
    suggested_base: str = ""
    coverage: str = ""
    why: str = ""
    candidates: str = ""
    existing: str = ""
    """Custom extraction views that already feed this object.

    Provenance like the rest, and the column people read first on a re-run.
    """

    row: int = 0
    """Sheet row number, so a complaint can name the line to fix."""

    @property
    def needs_base(self) -> bool:
        return self.action in (USE, WRAP)

    @property
    def generates(self) -> bool:
        """Does this row produce DDL? ``USE`` does not — that is its point."""
        return self.action in (WRAP, BUILD)

    @property
    def where(self) -> str:
        return f"row {self.row}" if self.row else self.object_name

    def problems(self) -> list[str]:
        """Everything wrong with this row, in the user's terms."""
        found: list[str] = []
        if not self.object_name:
            found.append(f"{self.where}: no object name")
        if self.action not in ACTIONS:
            found.append(
                f"{self.where}: Action {self.action!r} is not one of "
                f"{', '.join(ACTIONS)}"
            )
            # Nothing below can be judged without knowing the action.
            return found

        if self.needs_base and not self.base:
            found.append(
                f"{self.where}: {self.action} needs a view name in Base — "
                f"the Candidates column lists what was found"
            )
        if self.action == BUILD and self.base:
            found.append(
                f"{self.where}: BUILD generates over the table {self.object_name}, "
                f"so Base ({self.base}) would be ignored. Clear it, or use WRAP."
            )
        if self.generates and self.target and not _CUSTOMER_NAME.match(self.target):
            found.append(
                f"{self.where}: Target name {self.target!r} is not a "
                f"customer-namespace name — it must start with Z or Y"
            )
        return found


def validate(decisions: list[Decision]) -> list[str]:
    """Every problem across the whole sheet, before any of it runs.

    Whole-sheet rather than row-by-row because the failure that costs most is
    not a bad row — it is a batch that creates eleven objects and stops on a
    typo in the twelfth, leaving the user to work out which eleven.
    """
    problems: list[str] = []
    for decision in decisions:
        problems.extend(decision.problems())

    # Two rows generating the same name is not a typo either party would spot:
    # the second create fails on "already exists" and reads like a stale object.
    seen: dict[str, Decision] = {}
    for decision in decisions:
        if not decision.generates or not decision.target:
            continue
        key = decision.target.upper()
        if key in seen:
            problems.append(
                f"{decision.where}: Target name {decision.target} is already "
                f"used by {seen[key].object_name} ({seen[key].where}). Two rows "
                f"cannot generate the same object."
            )
        else:
            seen[key] = decision

    return problems


# ---------------------------------------------------------------------------
# Turning an assessment into a suggested row
# ---------------------------------------------------------------------------


def suggest(
    object_name: str,
    report,
    *,
    kind: str = "TABLE",
    estate=None,
    table=None,
) -> Decision:
    """Pre-fill one row from an F-09 successor report.

    The suggestion is a starting point and is written into *both* the editable
    ``Action`` column and the read-only ``Suggested action`` column. A user who
    changes nothing gets the tool's judgement; a user who changes something
    leaves a visible record of having disagreed.
    """
    decision = Decision(object_name=object_name, kind=kind)

    choices = report.choices() if hasattr(report, "choices") else []
    decision.candidates = " | ".join(_describe(c) for c in choices[:8])
    decision.why = getattr(report, "recommendation", "")

    best = getattr(report, "suggested", None)
    # The report decides USE / WRAP / BUILD; this only turns it into a row.
    # Re-deriving it here is what produced a WRAP row whose Why said "no
    # wrapper is needed".
    proposed = getattr(report, "action", None)
    if getattr(report, "readers_unknown", False):
        # The search could not run, which is not the same as finding nothing.
        # BUILD here would name a Z object for a table that may already have a
        # perfectly good view over it — the sheet has to say "unknown", and a
        # human has to look.
        decision.action = SKIP
        decision.note = "candidate search unavailable — see Why"
    elif getattr(report, "prefer_table", False) or best is None:
        # Ask the generator's own gate before proposing to generate. Measured
        # on twenty unseen tables: BSID and BSAK came back BUILD and `apply`
        # then refused both — they are S/4 compatibility views over ACDOCA, and
        # CDC needs a trigger on something real. The tool knew at plan time and
        # said it two steps later.
        refusal = _cannot_build(table)
        if refusal:
            decision.action = SKIP
            decision.note = "cannot be built on — see Why"
            # Replaces the recommendation rather than prefixing it. The report
            # ends "Build directly on the table", and keeping that after "this
            # cannot be built on" produces a cell that argues with itself.
            decision.why = (
                f"{refusal} Replicate the underlying table instead — in S/4 "
                f"these compatibility views read from ACDOCA or MATDOC, and "
                f"that is where the delta has to come from."
            )
        else:
            decision.action = BUILD
            decision.target = _default_target(object_name, BUILD)
    elif proposed == USE:
        # Already does the job. Generating anything here would be work for its
        # own sake, and a second object to keep in step with the first.
        decision.action = USE
        decision.base = best.view
        decision.coverage = f"{best.coverage:.0%}"
    else:
        decision.action = WRAP
        decision.base = best.view
        decision.coverage = f"{best.coverage:.0%}"
        decision.target = _default_target(best.view, WRAP)

    caveat = _release_caveat(best if decision.base else None)
    if caveat:
        decision.why = f"{decision.why} {caveat}".strip()

    if estate is not None:
        # What already exists, said before the action is fixed — a second run
        # over the same list otherwise proposes rebuilding everything the
        # first one built.
        decision.existing = estate.note_for(object_name, decision.base)
        if decision.existing.startswith("ALREADY BUILT:") and decision.generates:
            # An object over exactly this base already exists. Suggest leaving
            # it alone; the user can still override, and the original
            # suggestion is preserved below so the disagreement is visible.
            decision.action = SKIP
            decision.note = "already built — see the Existing column"

    decision.suggested_action = decision.action
    decision.suggested_base = decision.base
    return decision


def _cannot_build(table) -> str:
    """The generator's refusal for this table, or empty.

    ``None`` for the table means the caller did not look it up, which is not
    evidence of anything — so nothing is claimed and the suggestion stands as
    it was.
    """
    if table is None:
        return ""
    from cdcforge.generator.ztable import refuse_reason

    return refuse_reason(table)


def _describe(candidate) -> str:
    """One candidate, with the two facts coverage alone hides.

    ``FNDEI_EKPO_FILTER 85%`` outranks the chosen ``R_PURCHASINGDOCUMENTITEM
    54%`` on the only number shown, and is a worse answer: it is filtered, so
    it gives every column of *some* rows. A sheet that prints coverage and
    nothing else actively argues for the wrong choice.
    """
    marks = ""
    if getattr(candidate, "carries_cdc", False):
        marks += "*"
    if getattr(candidate, "row_filtered", False):
        marks += "~"
    return f"{candidate.view} {candidate.coverage:.0%}{marks}"


def _release_caveat(candidate) -> str:
    """Say it out loud when the chosen base is not a released API.

    ``VC_INTEGRATION_VBAP`` is SAP's own extraction view for VBAP — unfiltered,
    CDC-enabled, 78% coverage against 41% for the next best — and its own
    source explains why it carries no contract::

        // Unfortunately only views with a prefix 'I_' and 'C_' are allowed
        // to be released for C1 contract

    It is still the right answer, and it is still something SAP may change in
    a support pack without telling anyone. Both halves belong in the sheet,
    because the sheet is what leaves the building.
    """
    state = (getattr(candidate, "api_state", "") or "").upper()
    if not candidate or state in ("C1", "RELEASED"):
        return ""
    if state in ("", "UNKNOWN"):
        return (
            "Its release contract could not be read, so treat it as "
            "unreleased: check it after every upgrade."
        )
    return (
        "Note it is not a released API — SAP may change or withdraw it in a "
        "support pack without notice, so re-check it after every upgrade."
    )


def _default_target(source: str, action: str) -> str:
    """A Z name that is short enough to be legal and obvious enough to read.

    DDLS names cap at 30 characters. Truncating from the right keeps the part
    people recognise — ``ZW_SALESDOCUMENTITEMBASIC`` is still readable cut
    short, whereas a hash is not.
    """
    stem = re.sub(r"^[A-Z]_", "", (source or "").upper())
    stem = re.sub(r"[^A-Z0-9_]", "", stem)
    prefix = "ZW_" if action == WRAP else "ZI_"
    return (prefix + stem)[:30]


# ---------------------------------------------------------------------------
# Writing the sheet
# ---------------------------------------------------------------------------


def write_plan(decisions: list[Decision], path: str | Path) -> Path:
    """Write the decision workbook."""
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Alignment, Font, PatternFill
        from openpyxl.utils import get_column_letter
        from openpyxl.worksheet.datavalidation import DataValidation
    except ImportError as exc:  # pragma: no cover - environment dependent
        raise RuntimeError(
            "The decision sheet needs openpyxl: python -m pip install openpyxl."
        ) from exc

    target = Path(path)
    target.parent.mkdir(parents=True, exist_ok=True)

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Decisions"

    sheet.append(list(_COLUMNS))
    header_font = Font(bold=True, color="FFFFFF")
    editable_fill = PatternFill("solid", fgColor="1F5C34")
    fixed_fill = PatternFill("solid", fgColor="333333")
    for cell in sheet[1]:
        cell.font = header_font
        cell.fill = editable_fill if cell.value in _EDITABLE else fixed_fill

    for decision in decisions:
        sheet.append(
            [
                decision.object_name,
                decision.kind,
                decision.action,
                decision.base,
                decision.target,
                decision.note,
                decision.existing,
                decision.suggested_action,
                decision.suggested_base,
                decision.coverage,
                decision.why,
                decision.candidates,
            ]
        )

    # A dropdown rather than free text. Excel's own validation catches the
    # typo before the file is ever handed back, which is the only point at
    # which it is cheap to catch.
    if decisions:
        validation = DataValidation(
            type="list",
            formula1=f'"{",".join(ACTIONS)}"',
            allow_blank=False,
            showDropDown=False,
        )
        validation.error = "Pick one of USE, WRAP, BUILD or SKIP."
        validation.errorTitle = "Not an action"
        sheet.add_data_validation(validation)
        validation.add(f"C2:C{len(decisions) + 1}")

    sheet.freeze_panes = "C2"
    sheet.auto_filter.ref = sheet.dimensions
    widths = {
        "Object": 26, "Kind": 8, "Action": 10, "Base": 30, "Target name": 30,
        "Note": 28, "Existing": 54, "Suggested action": 16,
        "Suggested base": 30, "Coverage": 10, "Why": 60, "Candidates": 70,
    }
    for index, name in enumerate(_COLUMNS, start=1):
        sheet.column_dimensions[get_column_letter(index)].width = widths[name]
    wrapped = {_COLUMNS.index(n) + 1 for n in ("Existing", "Why", "Candidates")}
    for row in sheet.iter_rows(min_row=2):
        for cell in row:
            if cell.column in wrapped:
                cell.alignment = Alignment(wrap_text=True, vertical="top")

    _add_instructions(workbook)
    workbook.save(target)
    return target


def _add_instructions(workbook) -> None:
    """A sheet saying what the columns mean.

    Worth the space. This file gets mailed to someone who was not in the
    conversation where it was generated, and a column called ``Base`` means
    nothing on its own.
    """
    from openpyxl.styles import Alignment, Font

    sheet = workbook.create_sheet("How to use this")
    sheet["A1"] = "CDC Forge — decision sheet"
    sheet["A1"].font = Font(bold=True, size=14)

    lines = [
        ("", ""),
        ("Edit these columns", "Everything else is what the tool found, kept "
                               "so the reasoning survives."),
        ("  Action", "USE, WRAP, BUILD or SKIP — see below."),
        ("  Base", "The existing view to use or wrap. Pick from Candidates."),
        ("  Target name", "The Z object to generate. Must start with Z or Y."),
        ("  Note", "Yours. Carried through, never interpreted."),
        ("", ""),
        ("Existing", "Custom extraction views that already feed this object. "
                     "Read this column first on a re-run — it is what stops "
                     "the same object being built twice. ALREADY BUILT means "
                     "something exists over the very same base, and the "
                     "suggestion is SKIP."),
        ("", ""),
        ("Actions", ""),
    ]
    lines += [(f"  {name}", ACTION_HELP[name]) for name in ACTIONS]
    lines += [
        ("", ""),
        ("Candidates", "Existing views over the object, best first, with "
                       "column coverage."),
        ("  *", "Already carries extraction and CDC annotations."),
        ("  ~", "Row-filtered — a WHERE clause somewhere in its stack. It "
                "gives every column of SOME rows. Do not read its coverage "
                "as comparable: a filtered view at 85% is usually a worse "
                "answer than an unfiltered one at 54%."),
        ("Coverage", "How much of the table's columns the suggested view "
                     "exposes. It counts columns, not rows — which is exactly "
                     "why the ~ mark matters."),
        ("", ""),
        ("Then", "cdcforge apply --profile <id> --file <this file> "
                 "--out-dir ddl/"),
        ("", "Nothing is created until you add --create. The sheet is checked "
             "as a whole first, so a bad row stops the batch before any of it "
             "runs."),
    ]
    for label, text in lines:
        sheet.append([label, text])

    sheet.column_dimensions["A"].width = 20
    sheet.column_dimensions["B"].width = 96
    for row in sheet.iter_rows(min_col=2, max_col=2):
        for cell in row:
            cell.alignment = Alignment(wrap_text=True, vertical="top")


# ---------------------------------------------------------------------------
# Reading it back
# ---------------------------------------------------------------------------


def read_plan(source: str | Path | bytes) -> list[Decision]:
    """Read a decision sheet. Raises :class:`PlanError` if it is not one.

    Columns are found by *name*, not by position. A user who inserts a column
    for their own notes — which is the first thing anyone does to a spreadsheet
    — must not silently shift every value one place and have the tool build
    the wrong objects.
    """
    try:
        from openpyxl import load_workbook
    except ImportError as exc:  # pragma: no cover - environment dependent
        raise RuntimeError(
            "Reading the decision sheet needs openpyxl: pip install openpyxl."
        ) from exc

    if isinstance(source, bytes):
        import io

        handle = io.BytesIO(source)
    else:
        handle = Path(source)
        if not handle.exists():
            raise PlanError(f"{handle} does not exist")

    workbook = load_workbook(handle, read_only=True, data_only=True)
    sheet = None
    for candidate in workbook.worksheets:
        if candidate.title.strip().lower() == "decisions":
            sheet = candidate
            break
    sheet = sheet or workbook.worksheets[0]

    rows = sheet.iter_rows(values_only=True)
    try:
        header = next(rows)
    except StopIteration:
        raise PlanError(f"{sheet.title!r} is empty") from None

    index = {
        str(name).strip().lower(): position
        for position, name in enumerate(header)
        if name is not None
    }
    for required in ("object", "action"):
        if required not in index:
            raise PlanError(
                f"{sheet.title!r} has no {required.title()!r} column — this "
                f"does not look like a decision sheet. Generate one with "
                f"`cdcforge plan`."
            )

    def value(row, name: str) -> str:
        position = index.get(name)
        if position is None or position >= len(row):
            return ""
        cell = row[position]
        return "" if cell is None else str(cell).strip()

    decisions: list[Decision] = []
    for number, row in enumerate(rows, start=2):
        if not row or not value(row, "object"):
            continue
        decisions.append(
            Decision(
                object_name=value(row, "object").upper(),
                kind=(value(row, "kind") or "TABLE").upper(),
                action=(value(row, "action") or SKIP).upper(),
                base=value(row, "base").upper(),
                target=value(row, "target name").upper(),
                note=value(row, "note"),
                existing=value(row, "existing"),
                suggested_action=value(row, "suggested action").upper(),
                suggested_base=value(row, "suggested base").upper(),
                coverage=value(row, "coverage"),
                why=value(row, "why"),
                candidates=value(row, "candidates"),
                row=number,
            )
        )
    return decisions


@dataclass
class PlanSummary:
    """What a sheet asks for, counted."""

    decisions: list[Decision] = field(default_factory=list)
    problems: list[str] = field(default_factory=list)

    @property
    def ok(self) -> bool:
        return not self.problems

    def by_action(self, action: str) -> list[Decision]:
        return [d for d in self.decisions if d.action == action]

    @property
    def changed(self) -> list[Decision]:
        """Rows where the user disagreed with the tool.

        Worth surfacing on its own: it is the most interesting thing in a
        returned sheet, and the thing a reviewer wants to look at first.
        """
        return [
            d
            for d in self.decisions
            if d.suggested_action and d.action != d.suggested_action
        ]

    def render(self) -> str:
        counts = ", ".join(
            f"{len(self.by_action(a))} {a}" for a in ACTIONS if self.by_action(a)
        )
        lines = [f"{len(self.decisions)} row(s): {counts or 'nothing to do'}"]
        if self.changed:
            lines.append(f"  {len(self.changed)} changed from the suggestion:")
            for decision in self.changed[:10]:
                lines.append(
                    f"    {decision.object_name}: "
                    f"{decision.suggested_action} → {decision.action}"
                )
        for problem in self.problems:
            lines.append(f"  PROBLEM {problem}")
        return "\n".join(lines)


def load(source: str | Path | bytes) -> PlanSummary:
    """Read and validate in one step — the way callers actually want it."""
    decisions = read_plan(source)
    return PlanSummary(decisions=decisions, problems=validate(decisions))
