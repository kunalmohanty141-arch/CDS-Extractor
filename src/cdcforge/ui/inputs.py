"""Reading the user's object lists.

Separated from ``app.py`` so it can be tested: a Streamlit script executes
``st.*`` calls at import time and cannot be imported by a test.

The forgiving bits here are deliberate. A consultant pastes a column out of a
mail or a spreadsheet the customer sent, and it arrives with quotes, trailing
commas, blank rows and a header. Refusing that would make the tool feel broken
when the user did nothing wrong.
"""

from __future__ import annotations

import io

#: Prefixes that identify a real object name rather than a column heading.
#: Deliberately broad — a wrong guess here drops a row the user asked about.
_HEADER_HINTS = (
    "view", "name", "object", "cds", "table", "ddl", "entity", "description",
)


def split_names(text: str) -> list[str]:
    """Parse a pasted list into object names, de-duplicated, order preserved."""
    seen: set[str] = set()
    names: list[str] = []
    for raw in (text or "").replace(",", "\n").replace(";", "\n").split("\n"):
        name = raw.strip().strip("\"'").strip()
        if not name:
            continue
        upper = name.upper()
        if upper in seen:
            continue
        seen.add(upper)
        names.append(upper)
    return names


def looks_like_header(value: str) -> bool:
    """Is this first row a column heading rather than an object?

    Only drops a row that says something like "View name". Anything that could
    plausibly be an object is kept — losing a row the user asked about is worse
    than carrying one junk row they can see and ignore.
    """
    cleaned = value.strip().strip("\"'").lower()
    if not cleaned:
        return False
    return any(hint in cleaned for hint in _HEADER_HINTS) and len(cleaned) < 40


def read_csv(data: bytes) -> str:
    """First column of a CSV, as newline-separated names."""
    text = data.decode("utf-8-sig", "replace")
    rows = [line.split(",")[0].strip() for line in text.splitlines()]
    rows = [r for r in rows if r]
    if rows and looks_like_header(rows[0]):
        rows = rows[1:]
    return "\n".join(rows)


def _sheet_role(title: str) -> str:
    """``"views"``, ``"tables"`` or ``""`` from a worksheet's name."""
    lowered = (title or "").lower()
    if "table" in lowered:
        return "tables"
    if "view" in lowered or "cds" in lowered:
        return "views"
    return ""


def read_workbook(data: bytes) -> tuple[str, str]:
    """``(views, tables)`` from an uploaded workbook.

    Sheet *names* decide when they say something — a tab called "Tables" means
    tables whichever position it is in, and getting that backwards would tag
    every object wrongly. Otherwise position decides, first sheet views and
    second tables, because the spreadsheet arrives from the customer and its
    tabs will be called whatever they were called.
    """
    from openpyxl import load_workbook

    workbook = load_workbook(io.BytesIO(data), read_only=True, data_only=True)
    sheets = workbook.worksheets

    def first_column(sheet) -> str:
        values: list[str] = []
        for row in sheet.iter_rows(values_only=True):
            if not row:
                continue
            cell = row[0]
            if cell is None or not str(cell).strip():
                continue
            values.append(str(cell).strip())
        if values and looks_like_header(values[0]):
            values = values[1:]
        return "\n".join(values)

    named = {_sheet_role(s.title): s for s in sheets if _sheet_role(s.title)}
    if "views" in named or "tables" in named:
        views = first_column(named["views"]) if "views" in named else ""
        tables = first_column(named["tables"]) if "tables" in named else ""
        return views, tables

    views = first_column(sheets[0]) if sheets else ""
    tables = first_column(sheets[1]) if len(sheets) > 1 else ""
    return views, tables


def read_upload(filename: str, data: bytes) -> tuple[str, str]:
    """``(views, tables)`` from an uploaded CSV or workbook.

    A CSV is one column with no second sheet to carry tables, so everything
    lands in views and :func:`sort_by_kind` moves whatever is really a table.
    """
    if filename.lower().endswith(".csv"):
        return read_csv(data), ""
    return read_workbook(data)


def sort_by_kind(
    metadata, views: list[str], tables: list[str]
) -> tuple[list[str], list[str], list[tuple[str, str]], list[str]]:
    """Re-sort names by what the system says they are.

    Returns ``(views, tables, moved, unknown)`` where ``moved`` is
    ``(name, destination)`` pairs.

    Which sheet a name arrived on is the customer's opinion; whether DDIC holds
    a table or a DDL source of that name is a fact. Trusting the sheet meant a
    table listed under "views" was sent to the view validator, which reported
    it UNPARSEABLE — a confusing answer to a question the user never asked.

    Names the system does not recognise are returned separately rather than
    guessed at. They are usually a typo or an object in another client, and
    silently dropping them would hide that.
    """
    resolved_views: list[str] = []
    resolved_tables: list[str] = []
    moved: list[tuple[str, str]] = []
    unknown: list[str] = []

    for name, arrived_as in [(n, "views") for n in views] + [
        (n, "tables") for n in tables
    ]:
        is_table = metadata.get_table(name) is not None
        is_view = metadata.get_view_source(name) is not None

        # A CDS view entity has no DDIC table, and a table has no DDL source,
        # so the two are normally exclusive. When both answer — a classic view
        # and its generated SQL view can collide — the DDL source is the object
        # the user means.
        if is_view:
            destination = "views"
        elif is_table:
            destination = "tables"
        else:
            unknown.append(name)
            continue

        (resolved_views if destination == "views" else resolved_tables).append(name)
        if destination != arrived_as:
            moved.append((name, destination))

    return resolved_views, resolved_tables, moved, unknown
