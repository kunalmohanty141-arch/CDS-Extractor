"""F-34 — the assessment report.

    This is what makes the tool sellable to a consultancy. They run it in a day
    and hand the client a deliverable that would otherwise have taken three
    analysts six weeks. The report *is* the product for the assessment use case.

Excel, JSON and self-contained HTML. The HTML is the PDF path: a browser's
print-to-PDF produces a better document than reportlab would, and it avoids a
dependency whose only job would be layout.

A note on the effort estimate. F-34 asks for one, and an estimate assembled
from invented per-object numbers is exactly the false precision this tool
refuses elsewhere. So the model is explicit, configurable, and printed in the
report next to the number it produced. A reader can disagree with the
assumptions because they can see them.
"""

from __future__ import annotations

import html
import json
from dataclasses import dataclass, field
from datetime import datetime, timezone
from pathlib import Path

from cdcforge.model import Assessment
from cdcforge.store import Store
from cdcforge.triage import Bucket


@dataclass
class EffortModel:
    """Planning defaults, in person-days per object.

    These are not measurements. They are starting points a consultant should
    replace with their own rates, and the report says so wherever the total
    appears.
    """

    fixable_annotate_days: float = 0.5
    """A customer view that needs annotations added and transporting."""

    wrapper_days: float = 1.0
    """A released SAP view that needs a Z-wrapper built, tested and transported."""

    review_days: float = 1.0
    """An object a human has to look at before anything can be decided."""

    bare_table_days: float = 0.5
    """A Z-table with no view: generate, review, transport."""

    redesign_days: float = 3.0
    """A structurally impossible view. Aggregation, union or a to-many join
    means remodelling, not annotating — and that is a project, not a task."""

    def assumptions(self) -> list[str]:
        return [
            f"Fixable view (annotate in place): {self.fixable_annotate_days} days",
            f"Released SAP view (Z-wrapper): {self.wrapper_days} days",
            f"Manual review item: {self.review_days} days",
            f"Bare Z-table (generate a view): {self.bare_table_days} days",
            f"Structurally impossible (remodel): {self.redesign_days} days",
            "These are planning defaults, not measurements. Replace them with "
            "your own rates before quoting anything.",
        ]


@dataclass
class ReportData:
    """Everything the writers need, assembled once."""

    profile_id: str = "local"
    system_id: str = ""
    source_label: str = ""
    generated_at: str = ""
    counts: dict[str, int] = field(default_factory=dict)
    views: list[dict] = field(default_factory=list)
    tables: list[dict] = field(default_factory=list)
    bare_tables: list[dict] = field(default_factory=list)
    findings: list[dict] = field(default_factory=list)
    effort: EffortModel = field(default_factory=EffortModel)
    warnings: list[str] = field(default_factory=list)

    # -- assembly ----------------------------------------------------------
    @classmethod
    def from_store(
        cls,
        store: Store,
        assessments: list[Assessment] | None = None,
        *,
        system_id: str = "",
        source_label: str = "",
        effort: EffortModel | None = None,
    ) -> "ReportData":
        data = cls(
            profile_id=store.profile_id,
            system_id=system_id,
            source_label=source_label,
            generated_at=datetime.now(timezone.utc).isoformat(timespec="seconds"),
            effort=effort or EffortModel(),
        )
        data.views = store.views()
        data.tables = store.tables()
        data.bare_tables = store.tables(bare_only=True)

        data.counts = {bucket.value: 0 for bucket in Bucket}
        for row in data.views:
            bucket = row.get("bucket") or Bucket.REVIEW.value
            data.counts[bucket] = data.counts.get(bucket, 0) + 1
        data.counts["TOTAL_VIEWS"] = len(data.views)
        data.counts["BARE_TABLES"] = len(data.bare_tables)
        data.counts["EXTRACTION_ENABLED"] = sum(
            1 for r in data.views if r.get("extraction_enabled")
        )
        data.counts["CDC_ENABLED"] = sum(
            1 for r in data.views if (r.get("cdc_type") or "none") != "none"
        )

        for assessment in assessments or []:
            for result in assessment.problems:
                data.findings.append(
                    {
                        "object": assessment.object_name,
                        "verdict": assessment.verdict.value,
                        "rule": result.rule_id,
                        "outcome": result.outcome.value,
                        "severity": result.severity.value,
                        "line": result.ref.line,
                        "node": result.node,
                        "message": result.message,
                        "remediation": result.remediation,
                        "sap_source": result.sap_source,
                    }
                )

        if not assessments:
            data.warnings.append(
                "Per-finding detail was not supplied, so the report carries "
                "counts and per-object verdicts only."
            )

        # The counts come from the inventory's view records; the findings come
        # from the assessments handed in. They are different tables and they
        # can disagree. A report whose headline says nothing was scanned while
        # carrying findings for dozens of objects is contradicting itself, and
        # saying so is cheaper than letting someone read the wrong number.
        if not data.views and data.findings:
            objects = len({row["object"] for row in data.findings})
            data.warnings.append(
                f"The counts below are zero because no inventory has been run "
                f"against this store, but {len(data.findings)} finding(s) for "
                f"{objects} object(s) are included. Run 'cdc-forge inventory' "
                f"first for the totals to mean anything."
            )
        return data

    # -- derived -----------------------------------------------------------
    @property
    def effort_days(self) -> dict[str, float]:
        wrappers = sum(
            1
            for row in self.views
            if row.get("bucket") == Bucket.FIXABLE.value
            and (row.get("api_state") or "") in ("C1", "C2")
        )
        fixable = max(self.counts.get(Bucket.FIXABLE.value, 0) - wrappers, 0)
        return {
            "Fixable views (annotate)": fixable * self.effort.fixable_annotate_days,
            "Released SAP views (wrapper)": wrappers * self.effort.wrapper_days,
            "Manual review": self.counts.get(Bucket.REVIEW.value, 0) * self.effort.review_days,
            "Bare Z-tables": self.counts.get("BARE_TABLES", 0) * self.effort.bare_table_days,
            "Structurally impossible (remodel)":
                self.counts.get(Bucket.NOT_POSSIBLE.value, 0) * self.effort.redesign_days,
        }

    @property
    def total_effort_days(self) -> float:
        return round(sum(self.effort_days.values()), 1)

    @property
    def hot_tables(self) -> list[dict]:
        return [t for t in self.tables if t.get("is_hot")]

    def recommended_actions(self) -> list[str]:
        actions: list[str] = []
        ready = self.counts.get(Bucket.READY.value, 0)
        fixable = self.counts.get(Bucket.FIXABLE.value, 0)
        review = self.counts.get(Bucket.REVIEW.value, 0)
        impossible = self.counts.get(Bucket.NOT_POSSIBLE.value, 0)
        bare = self.counts.get("BARE_TABLES", 0)

        if ready:
            actions.append(
                f"{ready} view(s) are already CDC-ready. Start replication with "
                f"these — no development is needed."
            )
        if fixable:
            actions.append(
                f"{fixable} view(s) fail only on things the tool can generate: "
                f"missing annotations, incomplete mappings, unexposed keys."
            )
        if review:
            actions.append(
                f"{review} view(s) could not be settled statically. The usual "
                f"cause is an unproven to-one join cardinality — ABAP does not "
                f"validate it at runtime, and a wrong declaration produces "
                f"duplicate delta records weeks later."
            )
        if impossible:
            actions.append(
                f"{impossible} view(s) cannot be CDC-enabled as they stand "
                f"(aggregation, union, to-many join, parameters or a table "
                f"function). These need remodelling, not annotating."
            )
        if bare:
            actions.append(f"{bare} table(s) have no CDS view reading them at all.")
        if self.hot_tables:
            names = ", ".join(sorted(t["table_name"] for t in self.hot_tables)[:8])
            actions.append(
                f"High-write tables in scope ({names}). CDC adds database "
                f"triggers to the live write path; SAP publishes no cap on how "
                f"many tables can safely be enabled, so stage the rollout and "
                f"measure."
            )
        actions.append(
            "Corroborate every verdict with SAP's own pre-check (checkruns) "
            "before acting. Where the two disagree, the system wins."
        )
        return actions


# ---------------------------------------------------------------------------
# Writers
# ---------------------------------------------------------------------------


def write_json(data: ReportData, path: str | Path) -> Path:
    target = Path(path)
    target.parent.mkdir(parents=True, exist_ok=True)
    payload = {
        "profile_id": data.profile_id,
        "system_id": data.system_id,
        "source": data.source_label,
        "generated_at": data.generated_at,
        "counts": data.counts,
        "effort_days": data.effort_days,
        "total_effort_days": data.total_effort_days,
        "effort_assumptions": data.effort.assumptions(),
        "recommended_actions": data.recommended_actions(),
        "views": data.views,
        "tables": data.tables,
        "findings": data.findings,
        "warnings": data.warnings,
    }
    target.write_text(json.dumps(payload, indent=2, default=str), encoding="utf-8")
    return target


_VIEW_COLUMNS = [
    ("ddl_name", "View"),
    ("bucket", "List"),
    ("verdict", "Verdict"),
    ("entity_type", "Kind"),
    ("extraction_enabled", "Extraction"),
    ("delta_method", "Delta"),
    ("cdc_type", "CDC"),
    ("owner", "Owner"),
    ("api_state", "API state"),
    ("package", "Package"),
    ("base_tables", "Base tables"),
]

_TABLE_COLUMNS = [
    ("table_name", "Table"),
    ("table_class", "Class"),
    ("delivery_class", "Delivery"),
    ("owner", "Owner"),
    ("key_field_count", "Key fields"),
    ("field_count", "Fields"),
    ("est_rows", "Est. rows"),
    ("is_hot", "High write"),
    ("has_view", "Has a view"),
]

_FINDING_COLUMNS = [
    ("object", "Object"),
    ("verdict", "Verdict"),
    ("rule", "Rule"),
    ("outcome", "Outcome"),
    ("severity", "Severity"),
    ("line", "Line"),
    ("node", "Where"),
    ("message", "Finding"),
    ("remediation", "What to do"),
    ("sap_source", "SAP source"),
]


def _cell(value) -> str:
    if isinstance(value, list):
        return ", ".join(str(v) for v in value)
    if isinstance(value, bool):
        return "yes" if value else "no"
    if value is None:
        return ""
    return value


def write_excel(data: ReportData, path: str | Path) -> Path:
    """Excel workbook: summary, the three lists, tables, findings."""
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Alignment, Font, PatternFill
        from openpyxl.utils import get_column_letter
    except ImportError as exc:  # pragma: no cover - environment dependent
        raise RuntimeError(
            "Excel export needs openpyxl: python -m pip install openpyxl. "
            "The JSON and HTML exports have no dependencies."
        ) from exc

    target = Path(path)
    target.parent.mkdir(parents=True, exist_ok=True)

    workbook = Workbook()
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="333333")

    def add_sheet(title: str, columns, rows, first: bool = False):
        sheet = workbook.active if first else workbook.create_sheet()
        sheet.title = title
        sheet.append([label for _, label in columns])
        for cell in sheet[1]:
            cell.font = header_font
            cell.fill = header_fill
        for row in rows:
            sheet.append([_cell(row.get(key)) for key, _ in columns])
        sheet.freeze_panes = "A2"
        sheet.auto_filter.ref = sheet.dimensions
        for index, (key, label) in enumerate(columns, start=1):
            longest = max(
                [len(str(label))] + [len(str(_cell(r.get(key)))) for r in rows[:400]] or [10]
            )
            sheet.column_dimensions[get_column_letter(index)].width = min(max(longest + 2, 10), 70)
        return sheet

    # -- summary -----------------------------------------------------------
    summary = workbook.active
    summary.title = "Summary"
    summary["A1"] = "CDC Forge — assessment report"
    summary["A1"].font = Font(bold=True, size=14)

    rows = [
        ("System", data.system_id or data.profile_id),
        ("Source", data.source_label),
        ("Generated", data.generated_at),
        ("", ""),
        ("Views assessed", data.counts.get("TOTAL_VIEWS", 0)),
        ("  Ready", data.counts.get(Bucket.READY.value, 0)),
        ("  Fixable", data.counts.get(Bucket.FIXABLE.value, 0)),
        ("  Needs review", data.counts.get(Bucket.REVIEW.value, 0)),
        ("  Not possible", data.counts.get(Bucket.NOT_POSSIBLE.value, 0)),
        ("  Unparseable", data.counts.get(Bucket.UNPARSEABLE.value, 0)),
        ("", ""),
        ("Extraction enabled", data.counts.get("EXTRACTION_ENABLED", 0)),
        ("CDC enabled", data.counts.get("CDC_ENABLED", 0)),
        ("Tables with no view", data.counts.get("BARE_TABLES", 0)),
        ("", ""),
    ]
    for label, value in rows:
        summary.append([label, value])

    summary.append(["Effort estimate (person-days)", ""])
    for label, value in data.effort_days.items():
        summary.append([f"  {label}", value])
    summary.append(["  TOTAL", data.total_effort_days])
    summary.append(["", ""])
    summary.append(["Assumptions behind the estimate", ""])
    for line in data.effort.assumptions():
        summary.append([f"  {line}", ""])
    summary.append(["", ""])
    summary.append(["Recommended actions", ""])
    for action in data.recommended_actions():
        summary.append([f"  {action}", ""])

    summary.column_dimensions["A"].width = 46
    summary.column_dimensions["B"].width = 90
    for row in summary.iter_rows(min_col=2, max_col=2):
        for cell in row:
            cell.alignment = Alignment(wrap_text=True, vertical="top")

    ready = [v for v in data.views if v.get("bucket") == Bucket.READY.value]
    fixable = [v for v in data.views if v.get("bucket") == Bucket.FIXABLE.value]
    other = [
        v
        for v in data.views
        if v.get("bucket") not in (Bucket.READY.value, Bucket.FIXABLE.value)
    ]

    add_sheet("Ready", _VIEW_COLUMNS, ready)
    add_sheet("Fixable", _VIEW_COLUMNS, fixable)
    add_sheet("Review and blocked", _VIEW_COLUMNS, other)
    add_sheet("Bare tables", _TABLE_COLUMNS, data.bare_tables)
    add_sheet("All tables", _TABLE_COLUMNS, data.tables)
    if data.findings:
        add_sheet("Findings", _FINDING_COLUMNS, data.findings)

    workbook.save(target)
    return target


_HTML_STYLE = """
:root { color-scheme: light dark; }
* { box-sizing: border-box; }
body { font: 15px/1.55 -apple-system, "Segoe UI", Roboto, sans-serif;
       margin: 0; padding: 2.5rem; max-width: 1200px; margin-inline: auto;
       background: #fff; color: #1a1a1a; }
h1 { font-size: 1.7rem; margin: 0 0 .25rem; }
h2 { font-size: 1.15rem; margin: 2.5rem 0 .75rem; padding-bottom: .35rem;
     border-bottom: 2px solid #1a1a1a; }
.meta { color: #555; margin-bottom: 2rem; }
.cards { display: flex; flex-wrap: wrap; gap: .75rem; margin: 1rem 0 0; }
.card { border: 1px solid #d5d5d5; border-radius: 6px; padding: .75rem 1rem;
        min-width: 8.5rem; }
.card .n { font-size: 1.9rem; font-weight: 600; line-height: 1.1; }
.card .l { font-size: .8rem; text-transform: uppercase; letter-spacing: .04em;
           color: #555; }
.ready .n { color: #1a7f37; } .fixable .n { color: #9a6700; }
.review .n { color: #8250df; } .blocked .n { color: #cf222e; }
table { border-collapse: collapse; width: 100%; font-size: .86rem; margin-top: .5rem; }
th, td { border: 1px solid #ddd; padding: .35rem .5rem; text-align: left;
         vertical-align: top; }
th { background: #f2f2f2; position: sticky; top: 0; }
tbody tr:nth-child(even) { background: #fafafa; }
code { font: .85em ui-monospace, Consolas, monospace; }
ul { padding-left: 1.2rem; } li { margin: .3rem 0; }
.note { background: #f6f8fa; border-left: 4px solid #888; padding: .6rem .9rem;
        margin: 1rem 0; font-size: .9rem; }
.scroll { overflow-x: auto; }
@media print { body { padding: 0; } th { position: static; }
                h2 { break-after: avoid; } tr { break-inside: avoid; } }
"""


def _table_html(columns, rows, limit: int = 2000) -> str:
    if not rows:
        return "<p><em>Nothing in this list.</em></p>"
    head = "".join(f"<th>{html.escape(label)}</th>" for _, label in columns)
    body = []
    for row in rows[:limit]:
        cells = "".join(
            f"<td>{html.escape(str(_cell(row.get(key))))}</td>" for key, _ in columns
        )
        body.append(f"<tr>{cells}</tr>")
    more = (
        f"<p><em>{len(rows) - limit} more row(s) omitted; the Excel and JSON "
        f"exports carry the full list.</em></p>"
        if len(rows) > limit
        else ""
    )
    return (
        f'<div class="scroll"><table><thead><tr>{head}</tr></thead>'
        f"<tbody>{''.join(body)}</tbody></table></div>{more}"
    )


def write_html(data: ReportData, path: str | Path) -> Path:
    """Self-contained HTML. Print to PDF from a browser for the PDF deliverable."""
    target = Path(path)
    target.parent.mkdir(parents=True, exist_ok=True)

    def card(label: str, value, css: str = "") -> str:
        return (
            f'<div class="card {css}"><div class="n">{value}</div>'
            f'<div class="l">{html.escape(label)}</div></div>'
        )

    cards = "".join(
        [
            card("Ready", data.counts.get(Bucket.READY.value, 0), "ready"),
            card("Fixable", data.counts.get(Bucket.FIXABLE.value, 0), "fixable"),
            card("Needs review", data.counts.get(Bucket.REVIEW.value, 0), "review"),
            card("Not possible", data.counts.get(Bucket.NOT_POSSIBLE.value, 0), "blocked"),
            card("Unparseable", data.counts.get(Bucket.UNPARSEABLE.value, 0), "blocked"),
            card("Bare tables", data.counts.get("BARE_TABLES", 0)),
        ]
    )

    effort_rows = "".join(
        f"<tr><td>{html.escape(k)}</td><td>{v}</td></tr>"
        for k, v in data.effort_days.items()
    )
    assumptions = "".join(f"<li>{html.escape(a)}</li>" for a in data.effort.assumptions())
    actions = "".join(f"<li>{html.escape(a)}</li>" for a in data.recommended_actions())
    warnings = (
        "".join(f'<div class="note">{html.escape(w)}</div>' for w in data.warnings)
        if data.warnings
        else ""
    )

    ready = [v for v in data.views if v.get("bucket") == Bucket.READY.value]
    fixable = [v for v in data.views if v.get("bucket") == Bucket.FIXABLE.value]
    other = [
        v for v in data.views
        if v.get("bucket") not in (Bucket.READY.value, Bucket.FIXABLE.value)
    ]

    document = f"""<!doctype html>
<html lang="en"><head><meta charset="utf-8">
<meta name="viewport" content="width=device-width, initial-scale=1">
<title>CDC Forge — assessment report</title>
<style>{_HTML_STYLE}</style></head><body>
<h1>CDC readiness assessment</h1>
<div class="meta">
  System <strong>{html.escape(data.system_id or data.profile_id)}</strong> &middot;
  {html.escape(data.source_label)} &middot;
  generated {html.escape(data.generated_at)}
</div>
{warnings}
<div class="cards">{cards}</div>

<h2>What to do</h2>
<ul>{actions}</ul>

<h2>Effort estimate</h2>
<div class="scroll"><table><thead><tr><th>Work</th><th>Person-days</th></tr></thead>
<tbody>{effort_rows}
<tr><td><strong>Total</strong></td><td><strong>{data.total_effort_days}</strong></td></tr>
</tbody></table></div>
<div class="note">These figures come from the assumptions below, not from
measurement. They are a planning aid; replace the rates with your own before
quoting anything.</div>
<ul>{assumptions}</ul>

<h2>Ready &mdash; {len(ready)}</h2>
<p>Extraction and CDC enabled, and every deterministic rule satisfied.</p>
{_table_html(_VIEW_COLUMNS, ready)}

<h2>Fixable &mdash; {len(fixable)}</h2>
<p>A view exists but lacks extraction or delta, or its mapping is incomplete.
The tool can generate the corrected object.</p>
{_table_html(_VIEW_COLUMNS, fixable)}

<h2>Needs review or not possible &mdash; {len(other)}</h2>
<p>Either the rules could not decide, or the view is structurally incompatible
with CDC and needs remodelling rather than annotating.</p>
{_table_html(_VIEW_COLUMNS, other)}

<h2>Tables with no CDS view &mdash; {len(data.bare_tables)}</h2>
{_table_html(_TABLE_COLUMNS, data.bare_tables)}

<h2>Findings &mdash; {len(data.findings)}</h2>
{_table_html(_FINDING_COLUMNS, data.findings)}

<h2>How to read this</h2>
<p>Verdicts are deterministic where the rules allow, delegated to SAP where they
are not, and flagged for human review where neither is possible. A verdict of
<code>MANUAL_REVIEW</code> is a correct output, not a gap: the most common cause
is a declared to-one join cardinality that has never been checked against real
data. ABAP does not validate cardinality at runtime, so a wrong declaration
activates cleanly, loads correctly, and then produces duplicate or missing delta
records that nobody notices for weeks.</p>
<p>No business data was read to produce this report. Only metadata.</p>
</body></html>
"""
    target.write_text(document, encoding="utf-8")
    return target


def excel_available() -> bool:
    try:
        import openpyxl  # noqa: F401
    except ImportError:
        return False
    return True


def write_all(data: ReportData, directory: str | Path, stem: str = "cdc-assessment") -> list[Path]:
    """Write every available format.

    Excel is checked for *before* anything is written, not caught afterwards.
    Swallowing the failure at the end meant the JSON and HTML had already gone
    out with no mention that a format was missing — a report that quietly omits
    part of itself is the kind of silent gap this tool exists to complain about.
    """
    base = Path(directory)
    has_excel = excel_available()
    if not has_excel:
        data.warnings.append(
            "Excel export was skipped because openpyxl is not installed "
            "(pip install openpyxl). The JSON and HTML exports are complete."
        )

    written = [
        write_json(data, base / f"{stem}.json"),
        write_html(data, base / f"{stem}.html"),
    ]
    if has_excel:
        written.append(write_excel(data, base / f"{stem}.xlsx"))
    return written
